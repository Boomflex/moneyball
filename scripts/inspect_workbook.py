from __future__ import annotations

import json
import sys
from pathlib import Path

import openpyxl


def cell_style(cell):
    fill = cell.fill.fgColor.rgb if cell.fill and cell.fill.fgColor else None
    font = cell.font
    return {
        "fill": fill,
        "bold": bool(font and font.bold),
        "italic": bool(font and font.italic),
        "font_color": font.color.rgb if font and font.color and font.color.type == "rgb" else None,
        "num_fmt": cell.number_format,
        "alignment": {
            "horizontal": cell.alignment.horizontal,
            "vertical": cell.alignment.vertical,
            "wrap": cell.alignment.wrap_text,
        },
    }


def summarize_sheet(ws):
    rows = []
    formulas = []
    populated = []

    for row in ws.iter_rows():
        row_values = []
        has_value = False
        for cell in row:
            value = cell.value
            row_values.append(value)
            if value is not None:
                has_value = True
                populated.append(cell.coordinate)
                if isinstance(value, str) and value.startswith("="):
                    formulas.append({"cell": cell.coordinate, "formula": value})
        if has_value and len(rows) < 80:
            trimmed = list(row_values)
            while trimmed and trimmed[-1] is None:
                trimmed.pop()
            rows.append(trimmed[:40])

    validations = []
    for dv in ws.data_validations.dataValidation:
        validations.append(
            {
                "type": dv.type,
                "sqref": str(dv.sqref),
                "formula1": dv.formula1,
                "formula2": dv.formula2,
            }
        )

    style_samples = []
    for row in ws.iter_rows(min_row=1, max_row=min(ws.max_row, 30), min_col=1, max_col=min(ws.max_column, 20)):
        for cell in row:
            if cell.value is not None and len(style_samples) < 40:
                style_samples.append({"cell": cell.coordinate, "value": cell.value, "style": cell_style(cell)})

    return {
        "title": ws.title,
        "dimensions": ws.calculate_dimension(),
        "max_row": ws.max_row,
        "max_column": ws.max_column,
        "merged_ranges": [str(r) for r in ws.merged_cells.ranges],
        "freeze_panes": str(ws.freeze_panes) if ws.freeze_panes else None,
        "tables": list(ws.tables.keys()),
        "validations": validations,
        "rows_preview": rows,
        "formula_count": len(formulas),
        "formula_preview": formulas[:80],
        "style_samples": style_samples,
        "populated_count": len(populated),
    }


def main() -> None:
    if len(sys.argv) != 2:
        raise SystemExit("usage: inspect_workbook.py <workbook.xlsx>")

    workbook_path = Path(sys.argv[1])
    wb = openpyxl.load_workbook(workbook_path, data_only=False)
    summary = {
        "path": str(workbook_path),
        "sheets": [summarize_sheet(ws) for ws in wb.worksheets],
        "defined_names": sorted(name for name in wb.defined_names.keys()),
    }
    print(json.dumps(summary, indent=2, default=str))


if __name__ == "__main__":
    main()
