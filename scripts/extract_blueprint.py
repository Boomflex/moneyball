from __future__ import annotations

import json
import re
from pathlib import Path

import openpyxl


WORKBOOK_PATH = Path(r"C:/Users/jakek/Downloads/Moneyball Recruitment Spreadsheet 3 E14 Version.xlsx")
OUT_PATH = Path("src/model.js")

ROLE_SHEETS = {
    "GK": "GK Leagues",
    "CB": "CB Leagues",
    "FB": "FB Leagues",
    "MID": "MID Leagues",
    "Winger": "Winger Leagues",
    "Striker": "Striker Leagues",
}

SCORE_HEADERS = {
    "GK": ["Performance (z)"],
    "CB": ["Ball-Playing Score", "Stopper Score"],
    "FB": ["Defensive FB Score", "Attacking WB Score"],
    "MID": ["CDM Score", "CM Score", "CAM Score"],
    "Winger": ["Touchline Winger Score", "Inside Forward Score"],
    "Striker": ["Poacher Score", "Target Man Score", "False 9 / Creator Score"],
}

ROLE_ALIASES = {
    "GK": ["GK", "Goalkeeper"],
    "CB": ["CB", "Centre Back", "Center Back", "DC"],
    "FB": ["FB", "Full Back", "Fullback", "Wing Back", "WB", "LB", "RB", "DL", "DR"],
    "MID": ["DM", "CDM", "CM", "CAM", "AM", "MID", "Midfielder", "MC", "AMC", "DMC"],
    "Winger": ["Winger", "AML", "AMR", "ML", "MR", "Wide"],
    "Striker": ["ST", "Striker", "Forward", "CF"],
}


def col_name(index: int) -> str:
    name = ""
    while index:
        index, rem = divmod(index - 1, 26)
        name = chr(65 + rem) + name
    return name


def clean_header(value):
    return str(value).strip() if value is not None else ""


def parse_score_formula(formula: str, headers: dict[str, str]) -> tuple[list[dict], int]:
    formula = formula.replace("$", "")
    denom_match = re.search(r"\)/(\d+)(?:\+3|,\s*\"\"\))", formula)
    denominator = int(denom_match.group(1)) if denom_match else 100

    stats = []
    for match in re.finditer(r"(\d+)\*IFERROR\(\(([^()]+)\)/([0-9.]+),0\)", formula):
        weight = int(match.group(1))
        expression = match.group(2).strip()
        stdev = float(match.group(3))

        forward = re.match(r"([A-Z]+)2-([0-9.]+)", expression)
        reverse = re.match(r"([0-9.]+)-([A-Z]+)2", expression)
        if forward:
            col = forward.group(1)
            mean = float(forward.group(2))
            direction = 1
        elif reverse:
            mean = float(reverse.group(1))
            col = reverse.group(2)
            direction = -1
        else:
            continue

        stats.append(
            {
                "header": headers.get(col, col),
                "weight": weight,
                "mean": mean,
                "stdev": stdev,
                "direction": direction,
            }
        )
    return stats, denominator


def extract_leagues(ws):
    leagues = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        league = clean_header(row[0] if len(row) > 0 else "")
        if not league:
            continue
        def num(idx):
            value = row[idx] if len(row) > idx else None
            return value if isinstance(value, (int, float)) else None

        leagues[league] = {
            "strength": num(1),
            "valueScoreCoef": num(7),
            "valueAgeCoef": num(8),
            "valueIntercept": num(9),
            "wageScoreCoef": num(10),
            "wageAgeCoef": num(11),
            "wageIntercept": num(12),
        }
    return leagues


def extract_role(wb, sheet_name: str, league_sheet: str):
    ws = wb[sheet_name]
    headers_by_col = {col_name(c): clean_header(ws.cell(1, c).value) for c in range(1, ws.max_column + 1)}
    headers_by_index = {c: clean_header(ws.cell(1, c).value) for c in range(1, ws.max_column + 1)}

    league_col = next((c for c, h in headers_by_index.items() if h == "League Strength"), ws.max_column + 1)
    raw_headers = [h for c, h in headers_by_index.items() if c < league_col and h]

    scores = []
    for c, header in headers_by_index.items():
        if header not in SCORE_HEADERS[sheet_name]:
            continue
        formula = ws.cell(2, c).value
        if not isinstance(formula, str):
            continue
        stats, denominator = parse_score_formula(formula, headers_by_col)
        label = "GK Score" if sheet_name == "GK" and header == "Performance (z)" else header
        scores.append({"label": label, "stats": stats, "denominator": denominator})

    summary_values = {}
    for c, header in headers_by_index.items():
        value = ws.cell(2, c).value
        if header and isinstance(value, (int, float)):
            summary_values[header] = value

    deal_formula = ""
    for c, header in headers_by_index.items():
        if header == "Deal Flag":
            deal_formula = str(ws.cell(2, c).value or "")
            break
    threshold_match = re.search(r">=([0-9.]+)", deal_formula)

    return {
        "id": sheet_name,
        "label": sheet_name,
        "aliases": ROLE_ALIASES[sheet_name],
        "rawHeaders": raw_headers,
        "scoreColumns": scores,
        "leagues": extract_leagues(wb[league_sheet]),
        "summaryValues": summary_values,
        "freeAgentThreshold": float(threshold_match.group(1)) if threshold_match else None,
    }


def extract_archetypes(wb):
    ws = wb["Archetype Guide"]
    rows = []
    for r in range(5, ws.max_row + 1):
        role = clean_header(ws.cell(r, 2).value)
        archetype = clean_header(ws.cell(r, 3).value)
        meaning = clean_header(ws.cell(r, 4).value)
        metrics = clean_header(ws.cell(r, 5).value)
        if role and archetype:
            rows.append({"role": role, "archetype": archetype, "meaning": meaning, "metrics": metrics})
    return rows


def main() -> None:
    wb = openpyxl.load_workbook(WORKBOOK_PATH, data_only=False, read_only=False)
    model = {
        "sourceWorkbook": str(WORKBOOK_PATH),
        "roles": [extract_role(wb, role, league_sheet) for role, league_sheet in ROLE_SHEETS.items()],
        "archetypes": extract_archetypes(wb),
    }
    OUT_PATH.parent.mkdir(parents=True, exist_ok=True)
    OUT_PATH.write_text(
        "export const WORKBOOK_MODEL = "
        + json.dumps(model, indent=2, ensure_ascii=False)
        + ";\n",
        encoding="utf-8",
    )


if __name__ == "__main__":
    main()
